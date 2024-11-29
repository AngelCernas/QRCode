from openpyxl import load_workbook



consecutivo  = 92574
tarima = 2
area = 'TIMI'
bloque = '10B'
Px = 27
Py = '07'
Pz = '01'
bultos = 1
peso = 90
cubicaje = 0
marca = 'MRS ANA GARZA'
embalaje = 'CAJA'

parametros = {
    consecutivo,
    tarima,
    area,
    bloque,
    Px,
    Py,
    Pz,
    bultos,
    peso,
    cubicaje,
    marca,
    embalaje
}

file_path = 'formato.xlsx'

wb = load_workbook(file_path)

fila_init = 2
for parametros in parametros :
    file_path = 'formato.xlsx'

    fila_init += 2
    wb = load_workbook(file_path)
    print (parametros)
    ws = wb['Sheet1']
    wf = wb['Sheet1']
    ws['B2'] = parametros
    wf['C3'] = parametros

wb.save('file1.xlsx')
